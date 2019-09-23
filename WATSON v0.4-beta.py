from PIL import ImageGrab
import os
import time
import win32api, win32con
from PIL import ImageOps
from PIL import Image
from numpy import *
import math
import win32com.client
import openpyxl
from openpyxl import Workbook

    
#################mouse controls###################


def leftClick():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    time.sleep(1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0)

def rightClick():
    win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTDOWN,0,0)
    time.sleep(1)
    win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTUP,0,0)    

def doubleClick():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0)
    
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0)
    time.sleep(2)
    

def leftDown():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    time.sleep(.1)
    print('left down')

def leftUp():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    time.sleep(.1)
    print('left release')

def mousePos(cord):
    win32api.SetCursorPos((cord[0],cord[1]))

def get_cords():
    x,y=win32api.GetCursorPos()
    return x,y

###################level 2 functions########################
def set_screen(ent):
    count=0
    while True:
      if (count==500):
          return x,y
          break
      else:
         x,y = get_cords()
         print(x,y)
         time.sleep(0.02)
      count=count+1

def FindCursorPos():
    print("Take the cursor to the object on which to perform the action")
    time.sleep(2)
    curPos=input("All set! Press 1 to get position.\n")
    if(curPos=="1"):
     count=0
     while True:
       if (count==500):
           return x,y
           break
       else:
          x,y = get_cords()
          print(x,y)
          time.sleep(0.02)
       count=count+1
    

    

###############################computer vision #####################################################################
def Chk_grayscale_value(x1,y1,x2,y2):
   print("WATSON reading the greyscale avg. value...")
   time.sleep(1)
   count=1
   while True:
     if(count==200):
         break
     box=(x1,y1,x2,y2)
     im=ImageOps.grayscale(ImageGrab.grab(box))
     b=array(im.getcolors())
     b=b.sum()
     print(b)
     count= count+1  
   return b

def compare_xls_vision(x1,y1,x2,y2):
   box=(x1,y1,x2,y2)
   im=ImageOps.grayscale(ImageGrab.grab(box))
   b=array(im.getcolors())
   b=b.sum()
   print(b)
   return b

def leftclickWATSON(x,y):
    mousePos((x,y))
    leftClick()

def rightclickWATSON(x,y):
    mousePos((x,y))
    rightClick()    

def TypeWATSON(x,y,text):
    time.sleep(1)
    leftclickWATSON(x,y)
    time.sleep(1)
    shellb = win32com.client.Dispatch("WScript.Shell")
    shellb.SendKeys(text)

def JustTypeWATSON(text):
    shellb = win32com.client.Dispatch("WScript.Shell")
    shellb.SendKeys(text)
  
def doubleClickWATSON(x,y):
    mousePos((x,y))
    doubleClick()
#############################main funtction#####################################

def InitialSetup(integratorCall):
    Command_in=input("----------------------|OPTION|----------------------\n\nPress 1 to make WATSON see.\nPress 2 to assign a task to WATSON.\nPress 3 to save the process.\nPress 4 to automate the last saved process.\n\nwatson>>")
    Integrator(Command_in,1,integratorCall)
    
def Integrator(inp,ent,IC):
    curx1=0
    cury1=0
    curx2=0
    cury2=0
    if(inp=="1"):
       if(ent==1):
         print("Take the cursor to the top left corner of the view point")
         time.sleep(2)
         print("Keep it there for 10 seconds.")
         time.sleep(2)
         ready=input("press 1 and enter when you are ready...\n\n")
         if(ready=="1"):
          curx1,cury1= set_screen(ent)
          print("Take the cursor to the bottom right corner of the view point")
          time.sleep(2)
          print("Keep it there for 10 seconds.")
          time.sleep(2)
          ready=input("press 1 and enter when you are ready...\n\n")
          if(ready=="1"):
           curx2,cury2 = set_screen(2)
           time.sleep(1)
           print("Donot change the screen, WATSON reading the pixel values.")
           time.sleep(1)
           grayscaleVal=Chk_grayscale_value(curx1,cury1,curx2,cury2)
       print("Image rendering...")
       im = ImageGrab.grab((curx1,cury1,curx2,cury2))
       Vision="vision"
     #  curPos = str(curx1)+str(cury1)+str(curx2)+str(cury2)
       rows = (
         (Vision, curx1, cury1,curx2,cury2,grayscaleVal)

              )
       sheet.append(rows)
    
       IC = IC+1
       
       img_path = "full_snap_" + str(int(time.time())) + ".png"
       im.save(os.getcwd() + "\\full_snap_" + str(int(time.time())) + ".png", "PNG")
       image = Image.open(img_path)
       image.show()
       InitialSetup(IC)
       
    if(inp=="2"):
        ActionCmd = input("----------------------|ACTION|----------------------\n\n1.Right Click.\n2.Left Click.\n3.Just Type Here.\n4.Double Left Click .\n5.Key Press.\n6.Text input after click.\n7.Feed .xlsx data along the loop.\n8.Feed .xlsx data in new loop.\n\nwatson>>")
        cmd = "Action"+ str(ActionCmd)
        if(ActionCmd=="3"):
           textIn = input("Type text...\n\n")
           if(ActionCmd=="3"): 
            rows1 = (
                 (cmd,textIn)
                     )
        elif(ActionCmd=="6"):
           curx1, cury1=FindCursorPos() 
           textIn = input("Type text...\n\n")
           rows1 = (
                 (cmd,curx1, cury1,textIn)
                     )
               
        else:
            curx1, cury1=FindCursorPos()
            rows1 = (
                 (cmd, curx1, cury1)
                     )
        sheet.append(rows1)
        #sheet.append(IC, 1, str(curx1))
        #sheet.append(IC, 2, str(cury1))
    
        #workbook.close()
        print(curx1,cury1)
        IC = IC+1
        InitialSetup(IC)
    if(inp=="3"):
       FileName = input("Enter new project name\n\nwatson>>") 
       print("Process blueprint ready...")
       FileN=FileName+"_watson_dynamic.xlsx"
       book.save(FileN)
       IC = IC+1
       InitialSetup(IC)
    if(inp=="4"):
       FileName = input("Enter project name to be opened\n\nwatson>>") 
       FileN=FileName+"_watson_dynamic.xlsx"
       totalLoop = input("No. of loops.\n\nwatson>>")
       bookR = openpyxl.load_workbook(FileN)
       sheetR = bookR.active
       tl=0
       while(tl<int(totalLoop)):
        i=1
        while(i<=sheetR.max_row):
          
          commandWATSON = sheetR.cell(row=i, column=1)

          if(commandWATSON.value=="vision"):
            print("WATSON Vision - ")
            j=2
            attrArr=[]
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               attrArr.append(attr.value)
               j = j+1
            #####watson core module here##########
            while True: 
              gray_scale_val = compare_xls_vision(attrArr[0],attrArr[1],attrArr[2],attrArr[3])
              if(gray_scale_val==attrArr[4]):
                print("Watson can see now!")
                break
              print("vision unidentified.")   
            print("\n")

          if(commandWATSON.value=="Action1"):
            print("WATSON Action1 - ")
            j=2
            attrArr1=[]
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               attrArr1.append(attr.value)
               j = j+1
            rightclickWATSON(attrArr1[0],attrArr1[1])   
            print("\n")

          if(commandWATSON.value=="Action2"):
            print("WATSON Action2 - ")
            j=2
            attrArr2=[]
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               attrArr2.append(attr.value) 
               j = j+1
            leftclickWATSON(attrArr2[0],attrArr2[1])   
            print("\n")

          if(commandWATSON.value=="Action3"):
            print("WATSON Action3 - ")
            j=2
            attrArr3=[]
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               attrArr3.append(attr.value)
               j = j+1
            JustTypeWATSON(attrArr3[0])   
            print("\n")

          if(commandWATSON.value=="Action4"):
            print("WATSON Action4 - ")
            j=2
            attrArr4=[]
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               attrArr4.append(attr.value)
               j = j+1
            doubleClickWATSON(attrArr4[0],attrArr4[1])   
            print("\n")

          if(commandWATSON.value=="Action5"):
            print("WATSON Action5 - ")
            j=2  
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               print("attr : "+str(attr.value))
               j = j+1
            print("\n")

          if(commandWATSON.value=="Action6"):
            print("WATSON Action6 - ")
            j=2
            attrArr6=[]
            while(j<=sheetR.max_column):
               attr = sheetR.cell(row=i, column=j)
               if(attr.value==None):
                  break
               attrArr6.append(attr.value)
               j = j+1
            TypeWATSON(attrArr6[0],attrArr6[1],attrArr6[2])   
            print("\n")          
          i = i + 1
        tl = tl + 1
       IC = IC+1
       InitialSetup(IC)
######################main function#################


if __name__ == '__main__':
   book = Workbook()
   sheet = book.active
   curx1=0
   cury1=0
   curx2=0
   cury2=0
   integratorCall=1
   print("Hi! This is WATSON - Watching And Tracking Software Over Network.\n")
   time.sleep(2)
   print("Hyperpersonalised automation tool (HPAT)\n")
   time.sleep(2)
   print("Developed by - Shashank P. Sharma\n")
   time.sleep(2)
   AutoType = input("Press 1 for dynamically generated data automation.\nPress 2 for prefilled data automation(NOT INCLUDED IN THIS VERSION).\n\nwatson>>")
   if(AutoType=="1"):
      InitialSetup(integratorCall)
   else:
      InitialSetupPrefilled(integratorCall)       
   
